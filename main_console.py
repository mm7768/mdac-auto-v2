import imaplib
import email
import re
import base64
import random
import time
import io
import json
import sys
import os
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter import scrolledtext
from datetime import datetime
from queue import Queue

import ddddocr
from PIL import Image
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

import telebot
from filelock import FileLock

os.environ["PLAYWRIGHT_BROWSERS_PATH"] = "0"
CONFIG_FILE = "../dist/text/mdac_settings.json"
PROCESSED_EMAILS_FILE = "../dist/text/processed_emails.txt"  # 记录已处理的邮件ID

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath("../dist/text")
    return os.path.join(base_path, relative_path)

# ==========================================
# 线程安全的日志队列
# ==========================================
class LogQueue:
    def __init__(self):
        self.queue = Queue()

    def put(self, message, level="INFO", target_tab="MDAC"):
        self.queue.put((message, level, target_tab))

    def get(self):
        return self.queue.get()

    def empty(self):
        return self.queue.empty()

log_queue = LogQueue()

# ==========================================
# Excel 文件锁与状态管理
# ==========================================
class ExcelManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.lock = FileLock(f"{file_path}.lock", timeout=30)

    def update_status(self, row, new_status, remark=None):
        try:
            with self.lock:
                wb = load_workbook(self.file_path)
                sheet = wb.active
                sheet[f"I{row}"] = new_status
                sheet[f"J{row}"] = datetime.now()
                if remark is not None:
                    sheet[f"K{row}"] = remark
                wb.save(self.file_path)
        except Exception as e:
            log_queue.put(f"❌ Excel 状态更新失败 (行 {row}): {e}", level="ERROR", target_tab="MDAC")

    # ==========================================
    # Tab 3 专用：按护照号查找并写入 PIN 码
    # ==========================================
    def update_pin_by_passport(self, passport, pin):
        """
        在 Excel B 列查找护照号，找到后写入 PIN 到 H 列，
        并将 I 列状态改为 COMPLETED，J 列更新时间戳。
        返回: True=找到并写入, False=未找到
        """
        try:
            with self.lock:
                wb = load_workbook(self.file_path)
                sheet = wb.active
                passport_upper = passport.strip().upper()

                for row in range(2, sheet.max_row + 1):
                    cell_val = sheet[f"B{row}"].value
                    if cell_val and str(cell_val).strip().upper() == passport_upper:
                        sheet[f"H{row}"] = pin              # H列写入 PIN
                        sheet[f"I{row}"] = "COMPLETED"      # I列更新状态
                        sheet[f"J{row}"] = datetime.now()    # J列更新时间戳
                        wb.save(self.file_path)
                        return True
                return False
        except Exception as e:
            log_queue.put(f"❌ Excel 写入 PIN 失败: {e}", level="ERROR", target_tab="Gmail")
            return False


# ==========================================
# 2. 核心自动化逻辑 (Tab 1 MDAC)
# ==========================================

# ==========================================

def normalize_status(value):
    if value is None: return "PENDING"
    return str(value).strip().upper()


def read_customer_from_excel(sheet, row):
    def parse_date(val):
        if isinstance(val, datetime): return val
        if isinstance(val, str):
            val = val.strip()
            if not val: return None
            for fmt in ("%d-%b-%y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    return datetime.strptime(val.split()[0] if ' ' in val else val, fmt)
                except:
                    continue
            return None
        return None

    return {
        "name": str(sheet[f"A{row}"].value or "").strip(),
        "passport": str(sheet[f"B{row}"].value or "").strip(),
        "dob": parse_date(sheet[f"C{row}"].value),
        "sex_text": str(sheet[f"D{row}"].value or "").strip(),
        "passport_exp": parse_date(sheet[f"E{row}"].value),
        "arrdt": parse_date(sheet[f"F{row}"].value),
        "depdt": parse_date(sheet[f"G{row}"].value),
        "status": normalize_status(sheet[f"I{row}"].value),
        "last_check": sheet[f"J{row}"].value,
        "remark": str(sheet[f"K{row}"].value or "").strip(),
        "nationality": str(sheet[f"L{row}"].value or "CHN").strip().upper(),
    }


def validate_customer(customer):
    if not customer["name"]: raise ValueError("Name 不能为空")
    if not customer["passport"]: raise ValueError("Book Number 不能为空")
    if customer["dob"] is None: raise ValueError("Date of Birth 不能为空")
    if customer["passport_exp"] is None: raise ValueError("Date of Expiry 不能为空")
    if customer["arrdt"] is None: raise ValueError("Date of Arrival 不能为空")
    if customer["depdt"] is None: raise ValueError("Date of Departure 不能为空")
    if not customer["nationality"]: raise ValueError("Nationality 国籍不能空")

    if customer["sex_text"] == "男":
        return "1"
    elif customer["sex_text"] == "女":
        return "2"
    else:
        raise ValueError(f"未知性别：{customer['sex_text']}")


def process_registration(page, excel_manager, row, customer, config, log_func):
    try:
        sex = validate_customer(customer)
        name = customer["name"]
        passport = customer["passport"]
        dob = customer["dob"]
        passport_exp = customer["passport_exp"]
        arrdt = customer["arrdt"]
        depdt = customer["depdt"]
        nationality = customer["nationality"]

        excel_manager.update_status(row, "PROCESSING", "")
        log_func(f"第 {row} 行状态已更新：PROCESSING")

        dialog_messages = []

        def handle_dialog(dialog):
            message = dialog.message.strip()
            log_func(f"Dialog：{message}")
            dialog_messages.append(message)
            dialog.accept()

        page.on("dialog", handle_dialog)

        # =========================
        # 1. 进入页面
        # =========================
        page.goto("https://imigresen-online.imi.gov.my/mdac/main?registerMain", wait_until="domcontentloaded")
        page.wait_for_timeout(2000)

        # =========================
        # 2. 填写个人资料
        # =========================
        region = str(config.get("region", "60")).strip() or "60"
        page.locator("#region").select_option(region)
        page.locator("#nationality").select_option(nationality)
        # 修改点：Place of Birth 跟着 Nationality
        page.locator("#pob").select_option(nationality)

        page.locator("#email").fill(config.get("email", ""))
        page.locator("#confirmEmail").fill(config.get("email", ""))
        page.locator("#mobile").fill(config.get("phone", ""))
        page.locator("#trvlMode").select_option("2")
        page.locator("#embark").select_option("CHN")
        page.locator("#vesselNm").fill(config.get("vessel", ""))
        page.locator("#accommodationAddress1").fill(config.get("address1", ""))
        page.locator("#accommodationStay").select_option("02")
        page.locator("#accommodationAddress2").fill(config.get("address2", ""))
        page.locator("#accommodationState").select_option("01")
        page.locator("#accommodationCity").select_option("0100")
        page.locator("#accommodationPostcode").fill(config.get("postcode", ""))

        page.locator("#sex").select_option(sex)
        page.locator("#name").fill(name)
        page.locator("#passNo").fill(passport)

        # 使用 JS 强制填入日期
        dob_str = dob.strftime("%d/%m/%Y")
        exp_str = passport_exp.strftime("%d/%m/%Y")
        arrdt_str = arrdt.strftime("%d/%m/%Y")
        depdt_str = depdt.strftime("%d/%m/%Y")

        page.evaluate(f'''() => {{
            function setDateValue(id, value) {{
                let el = document.getElementById(id);
                if(el) {{
                    el.value = value;
                    el.dispatchEvent(new Event('input', {{ bubbles: true }}));
                    el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                }}
            }}
            setDateValue("dob", "{dob_str}");
            setDateValue("passExpDte", "{exp_str}");
            setDateValue("arrDt", "{arrdt_str}");
            setDateValue("depDt", "{depdt_str}");
        }}''')

        log_func("✅ 日期资料已通过 JS 极速填入")
        page.wait_for_timeout(500)

        # =========================
        # 3. 破解滑块 (完全保留用户原始逻辑)
        # =========================
        def solve_mdac_slider(page, max_retries=3):
            for attempt in range(max_retries):
                log_func(f"🔄 正在尝试第 {attempt + 1} 次滑块验证...")
                try:
                    page.wait_for_selector('canvas', timeout=10000)
                    page.wait_for_timeout(1500)

                    bg_base64 = page.evaluate("document.querySelectorAll('canvas')[0].toDataURL('image/png')")
                    block_base64 = page.evaluate("document.querySelectorAll('canvas')[1].toDataURL('image/png')")

                    bg_bytes = base64.b64decode(bg_base64.split(',')[1])
                    block_bytes = base64.b64decode(block_base64.split(',')[1])

                    block_img = Image.open(io.BytesIO(block_bytes))
                    bbox = block_img.getbbox()
                    img_start_x = bbox[0] if bbox else 0

                    det = ddddocr.DdddOcr(det=False, ocr=False, show_ad=False)
                    res = det.slide_match(block_bytes, bg_bytes, simple_target=True)

                    distance = res['target'][0] - img_start_x

                    scale_info = page.evaluate("""() => {
                        let cvs = document.querySelectorAll('canvas')[0];
                        return {
                            internal: cvs.width,
                            display: cvs.getBoundingClientRect().width
                        };
                    }""")
                    scale = scale_info['display'] / scale_info['internal'] if scale_info['internal'] else 1
                    final_distance = distance * scale

                    def generate_track(total_distance):
                        track = []
                        current = 0
                        steps = random.randint(30, 40)
                        for i in range(1, steps + 1):
                            progress = i / steps
                            ease_progress = 1 if progress == 1 else 1 - (2 ** (-10 * progress))
                            move = total_distance * ease_progress
                            step_move = move - current
                            current = move
                            track.append(step_move)
                        return track

                    slider_handle = page.locator('.slider').first
                    box = slider_handle.bounding_box()
                    handle_start_x = box['x'] + box['width'] / 2
                    handle_start_y = box['y'] + box['height'] / 2

                    slider_handle.hover()
                    page.mouse.down()
                    page.wait_for_timeout(random.randint(100, 200))

                    offset = -14
                    actual_move = final_distance + offset
                    track = generate_track(actual_move)
                    current_x = handle_start_x

                    for step_x in track:
                        current_x += step_x
                        page.mouse.move(current_x, handle_start_y + random.uniform(-1.5, 1.5))
                        time.sleep(random.uniform(0.01, 0.02))

                    page.wait_for_timeout(random.randint(300, 500))
                    page.mouse.up()

                    page.wait_for_timeout(2000)
                    success = page.evaluate(
                        "document.querySelector('.sliderContainer') !== null && document.querySelector('.sliderContainer').classList.contains('sliderContainer_success')"
                    )

                    if success:
                        return True
                    else:
                        log_func(f"⚠️ 第 {attempt + 1} 次验证失败，等待滑块重置...")
                        page.wait_for_timeout(2500)

                except Exception as e:
                    log_func(f"自动滑块破解发生错误: {e}")
                    page.mouse.up()
                    page.wait_for_timeout(2000)

            log_func("❌ 连续 3 次滑块验证失败！")
            return False

        slider_result = solve_mdac_slider(page)
        if slider_result:
            log_func("✅ 滑块验证成功！")
        else:
            log_func("❌ 滑块连续 3 次验证失败，已跳过当前客户！")
            excel_manager.update_status(row, "ERROR", "滑块连续3次验证失败")
            return

        page.wait_for_timeout(1000)

        # =========================
        # 4. 提交逻辑
        # =========================
        submit_button = page.get_by_role("button", name="Submit")
        submit_button.wait_for(state="visible", timeout=30000)

        if config.get("test_mode", True):
            log_func("\n当前为 TEST_MODE (测试模式)。不会点击 Submit。")
            excel_manager.update_status(row, "TESTED", "测试模式：未提交")
            log_func("请在浏览器中检查字段，5秒后将自动处理下一位客户...")
            page.wait_for_timeout(5000)
        else:
            log_func("准备点击 Submit。")
            submit_button.click()
            page.wait_for_timeout(8000)

            all_dialog_text = " ".join(dialog_messages).lower()
            success_found = (
                    "successful" in all_dialog_text or "successfully" in all_dialog_text or "success" in all_dialog_text or (
                    "pin" in all_dialog_text and "email" in all_dialog_text))

            if success_found:
                excel_manager.update_status(row, "REGISTERED", "")
                log_func("🎉 Registration 成功！")
            else:
                dialog_text = " | ".join(dialog_messages) or "没有捕获到任何 Dialog"
                excel_manager.update_status(row, "MANUAL_CHECK", dialog_text[:300])
                log_func("\n⚠️ 没有确认 Registration 成功，状态已更新为 MANUAL_CHECK。")

    except Exception as e:
        log_func(f"❌ 处理第 {row} 行客户时发生错误: {e}", level="ERROR")
        excel_manager.update_status(row, "ERROR", str(e)[:300])
    finally:
        page.remove_listener("dialog", handle_dialog)


# ==========================================
# 5. 新增：Gmail PIN 码自动获取模块 (Tab 3)
# ==========================================
class GmailPINFetcher:
    """
    定时从 Gmail 读取 mdac@imi.gov.my 发来的邮件，
    提取姓名、护照号、PIN 码，写入 Excel 或通过 Telegram 报警。
    """

    # 目标发件人（MDAC 官方）
    SENDER_FILTER = "mdac@imi.gov.my"

    def __init__(self, email_addr, app_password, excel_manager, telegram_token, chat_id,
                 interval_minutes, log_queue):
        self.email_addr = email_addr
        self.app_password = app_password
        self.excel_manager = excel_manager
        self.telegram_token = telegram_token
        self.chat_id = chat_id
        self.interval_minutes = max(1, interval_minutes)
        self.log_queue = log_queue
        self.running = False

    # ------------------------------------------------------------------
    # 本地记录：处理过的邮件 Message-ID
    # ------------------------------------------------------------------
    def _load_processed_ids(self):
        """从本地文件加载已处理的邮件 ID 集合"""
        try:
            if os.path.exists(PROCESSED_EMAILS_FILE):
                with open(PROCESSED_EMAILS_FILE, "r", encoding="utf-8") as f:
                    return set(line.strip() for line in f if line.strip())
        except Exception as e:
            self.log_queue.put(f"⚠️ 读取 processed_emails.txt 失败: {e}", level="WARNING", target_tab="Gmail")
        return set()

    def _save_processed_ids(self, processed_set):
        """将已处理的邮件 ID 写回本地文件"""
        try:
            os.makedirs(os.path.dirname(PROCESSED_EMAILS_FILE), exist_ok=True)
            with open(PROCESSED_EMAILS_FILE, "w", encoding="utf-8") as f:
                for mid in processed_set:
                    f.write(mid + "\n")
        except Exception as e:
            self.log_queue.put(f"⚠️ 保存 processed_emails.txt 失败: {e}", level="WARNING", target_tab="Gmail")

    # ------------------------------------------------------------------
    # 邮件解析
    # ------------------------------------------------------------------
    def _parse_email_body(self, msg):
        """从邮件中提取纯文本正文"""
        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                content_type = part.get_content_type()
                content_disposition = str(part.get("Content-Disposition", ""))
                # 跳过附件
                if "attachment" in content_disposition:
                    continue
                if content_type == "text/plain":
                    charset = part.get_content_charset() or "utf-8"
                    payload = part.get_payload(decode=True)
                    if payload:
                        body += payload.decode(charset, errors="ignore")
                elif content_type == "text/html":
                    # 尝试获取纯文本版本，没有则从 HTML 提取
                    charset = part.get_content_charset() or "utf-8"
                    payload = part.get_payload(decode=True)
                    if payload:
                        html_text = payload.decode(charset, errors="ignore")
                        # 简单去标签，只保留文字内容
                        import re as _re
                        body += _re.sub(r'<[^>]+>', ' ', html_text)
        else:
            charset = msg.get_content_charset() or "utf-8"
            payload = msg.get_payload(decode=True)
            if payload:
                body = payload.decode(charset, errors="ignore")
        return body

    def _extract_fields(self, body):
        """
        从邮件正文提取姓名、护照号、PIN 码。
        返回 dict，缺字段则对应值为 None。
        """
        name = None
        passport = None
        pin = None

        # Name : TANG FUMING
        m = re.search(r"Name\s*:\s*(.+?)(?:\r?\n|$)", body)
        if m:
            name = m.group(1).strip()

        # Passport No. : EJ1660876
        m = re.search(r"Passport\s+No\.\s*:\s*([A-Za-z0-9]+)", body)
        if m:
            passport = m.group(1).strip()

        # PIN : 8pczkJDr  (在 "Thank you" 之前，PIN 单独加粗显示)
        m = re.search(r"\bPIN\s*:\s*([A-Za-z0-9]+)", body)
        if m:
            pin = m.group(1).strip()

        return {"name": name, "passport": passport, "pin": pin}

    # ------------------------------------------------------------------
    # Telegram 报警
    # ------------------------------------------------------------------
    def _send_telegram_alert(self, name, passport, pin):
        """找不到护照号时通过 Telegram 推送报警"""
        if not self.telegram_token or not self.chat_id:
            self.log_queue.put("⚠️ Telegram Token 或 Chat ID 未配置，无法发送报警。", level="WARNING", target_tab="Gmail")
            return
        try:
            bot = telebot.TeleBot(self.telegram_token)
            text = (
                f"⚠️ **PIN码匹配失败（Excel 找不到护照号）**\n\n"
                f"客户名字：{name}\n"
                f"护照号：{passport}\n"
                f"PIN码：{pin}\n\n"
                f"请手动核查 Excel 文件。"
            )
            bot.send_message(self.chat_id, text, parse_mode="Markdown")
            self.log_queue.put(f"📨 Telegram 报警已发送: 护照号 {passport}", target_tab="Gmail")
        except Exception as e:
            self.log_queue.put(f"❌ Telegram 报警发送失败: {e}", level="ERROR", target_tab="Gmail")

    # ------------------------------------------------------------------
    # 单次执行：连接 Gmail → 提取 → 写入/报警
    # ------------------------------------------------------------------
    def _fetch_once(self):
        """执行一轮 Gmail 检查"""
        self.log_queue.put(f"[Gmail] 开始连接 Gmail 检查新邮件...", target_tab="Gmail")

        processed_ids = self._load_processed_ids()

        try:
            mail = imaplib.IMAP4_SSL("imap.gmail.com")
            mail.login(self.email_addr, self.app_password)
            mail.select("inbox")

            # 搜索来自 MDAC 的所有邮件（不分已读未读）
            status, messages = mail.search(None, f'(FROM "{self.SENDER_FILTER}")')
            if status != "OK":
                self.log_queue.put("⚠️ IMAP 搜索失败。", level="WARNING", target_tab="Gmail")
                mail.logout()
                return

            email_ids = messages[0].split()
            self.log_queue.put(f"[Gmail] 共找到 {len(email_ids)} 封来自 {self.SENDER_FILTER} 的邮件", target_tab="Gmail")

            new_count = 0
            for e_id in email_ids:
                # 获取 Message-ID 用于去重
                res, header_data = mail.fetch(e_id, "(BODY[HEADER.FIELDS (MESSAGE-ID)])")
                header_text = header_data[0][1].decode("utf-8", errors="ignore")
                msg_id_match = re.search(r"Message-ID:\s*<(.+?)>", header_text)
                if not msg_id_match:
                    continue
                message_id = msg_id_match.group(1).strip()

                # 跳过已处理
                if message_id in processed_ids:
                    continue

                # 获取完整邮件
                res, msg_data = mail.fetch(e_id, "(RFC822)")
                raw_email = msg_data[0][1]
                msg = email.message_from_bytes(raw_email)

                # 解析正文
                body = self._parse_email_body(msg)
                fields = self._extract_fields(body)

                if not fields["pin"]:
                    self.log_queue.put(f"⚠️ 邮件 Message-ID={message_id} 未找到 PIN 码，跳过。", level="WARNING", target_tab="Gmail")
                    processed_ids.add(message_id)  # 即使没 PIN 也标记已处理，避免无限重试
                    continue

                new_count += 1
                self.log_queue.put(
                    f"[Gmail] 提取到新邮件: 姓名={fields['name']}, 护照号={fields['passport']}, PIN={fields['pin']}",
                    target_tab="Gmail"
                )

                # 在 Excel 中查找护照号
                passport = fields["passport"]
                if passport:
                    found = self.excel_manager.update_pin_by_passport(passport, fields["pin"])
                    if found:
                        self.log_queue.put(
                            f"✅ 成功写入 PIN: {fields['name']} ({passport}) → {fields['pin']}",
                            target_tab="Gmail"
                        )
                        processed_ids.add(message_id)
                    else:
                        self.log_queue.put(
                            f"❌ 护照号 {passport} 在 Excel 中未找到，触发 Telegram 报警",
                            level="ERROR", target_tab="Gmail"
                        )
                        self._send_telegram_alert(
                            fields.get("name", "未知"),
                            passport,
                            fields["pin"]
                        )
                        # 不标记为已处理，下一轮会重试
                else:
                    self.log_queue.put(f"⚠️ 邮件未提取到护照号，跳过: Message-ID={message_id}", level="WARNING", target_tab="Gmail")
                    processed_ids.add(message_id)

            self._save_processed_ids(processed_ids)

            if new_count == 0:
                self.log_queue.put(f"[Gmail] 本轮无新邮件，等待 {self.interval_minutes} 分钟后重试...", target_tab="Gmail")

            mail.logout()

        except imaplib.IMAP4.error as e:
            self.log_queue.put(f"❌ Gmail 登录失败: {e}，请检查邮箱和密码。", level="ERROR", target_tab="Gmail")
        except Exception as e:
            self.log_queue.put(f"❌ Gmail 检查发生错误: {e}", level="ERROR", target_tab="Gmail")

    # ------------------------------------------------------------------
    # 线程控制
    # ------------------------------------------------------------------
    def start(self):
        """启动后台循环线程"""
        self.running = True
        self._thread = threading.Thread(target=self._run_loop, daemon=True)
        self._thread.start()
        self.log_queue.put(
            f"Gmail PIN 获取线程已启动，每 {self.interval_minutes} 分钟检查一次。",
            target_tab="Gmail"
        )

    def stop(self):
        """停止后台循环线程"""
        self.running = False
        self.log_queue.put("Gmail PIN 获取线程正在停止...", target_tab="Gmail")

    def _run_loop(self):
        """主循环：定时执行 fetch_once"""
        while self.running:
            self._fetch_once()
            # 可中断的等待
            for i in range(self.interval_minutes * 60):
                if not self.running:
                    break
                time.sleep(1)
        self.log_queue.put("Gmail PIN 获取线程已停止。", target_tab="Gmail")


# ==========================================
# 6. GUI 界面类
# ==========================================
class MDACApp:
    def __init__(self, root):
        self.root = root
        self.root.title("MDAC 全自动中央控制台")

        try:
            self.root.iconbitmap(resource_path("logo.ico"))
        except Exception as e:
            print(f"图标加载失败：{e}")

        self.root.geometry("800x750")
        self.root.configure(padx=10, pady=10)

        self.config = self.load_config()
        self.excel_manager = None
        self.is_mdac_running = False
        self.is_mdac_paused = False
        self.mdac_thread = None

        # Tab 3 状态
        self.is_gmail_running = False
        self.gmail_fetcher = None

        self.create_widgets()
        self.process_log_queue()

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except:
                pass
        return {
            "excel_path": "", "email": "", "phone": "", "region": "60", "vessel": "",
            "address1": "", "address2": "", "postcode": "", "test_mode": True,
            # Tab 3 Gmail 配置
            "gmail_address": "", "gmail_app_password": "",
            "gmail_interval": 10,
            "gmail_telegram_token": "", "gmail_chat_id": "",
        }

    def save_config(self):
        self.config = {
            "excel_path": self.excel_var.get(),
            "email": self.email_var.get(),
            "phone": self.phone_var.get(),
            "region": self.region_var.get().strip() or "60",
            "vessel": self.vessel_var.get(),
            "address1": self.addr1_var.get(),
            "address2": self.addr2_var.get(),
            "postcode": self.postcode_var.get(),
            "test_mode": self.test_mode_var.get(),
            # Tab 3 Gmail 配置
            "gmail_address": self.gmail_address_var.get(),
            "gmail_app_password": self.gmail_app_password_var.get(),
            "gmail_interval": int(self.gmail_interval_var.get()) if self.gmail_interval_var.get() else 10,
            "gmail_telegram_token": self.gmail_telegram_token_var.get(),
            "gmail_chat_id": self.gmail_chat_id_var.get(),
        }
        os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(self.config, f, ensure_ascii=False, indent=4)

    def create_widgets(self):
        style = ttk.Style()
        style.configure("TLabel", font=("微软雅黑", 10))
        style.configure("TButton", font=("微软雅黑", 10, "bold"))
        style.configure("TNotebook.Tab", font=("微软雅黑", 10, "bold"))

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill="both", padx=5, pady=5)

        # Tab 1: MDAC 控制
        self.mdac_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.mdac_tab, text="MDAC 控制")

        mdac_config_frame = ttk.LabelFrame(self.mdac_tab, text=" MDAC 固定资料设置 ", padding=15)
        mdac_config_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        def add_mdac_config_row(parent, label_text, var_name, row, is_file=False):
            ttk.Label(parent, text=label_text).grid(row=row, column=0, sticky=tk.W, pady=2)
            var = tk.StringVar(value=self.config.get(var_name, ""))
            setattr(self, f"{var_name}_var", var)
            entry = ttk.Entry(parent, textvariable=var, width=40)
            entry.grid(row=row, column=1, sticky=tk.W, padx=10, pady=2)

            if is_file:
                btn = ttk.Button(parent, text="浏览...", command=lambda: self.browse_excel_file(var))
                btn.grid(row=row, column=2, sticky=tk.W, padx=5, pady=2)
            return var

        self.excel_var = add_mdac_config_row(mdac_config_frame, "Excel 文件路径:", "excel_path", 0, is_file=True)
        self.email_var = add_mdac_config_row(mdac_config_frame, "Email:", "email", 1)
        self.phone_var = add_mdac_config_row(mdac_config_frame, "手机号:", "phone", 2)
        self.region_var = add_mdac_config_row(mdac_config_frame, "区域代码:", "region", 3)
        self.vessel_var = add_mdac_config_row(mdac_config_frame, "航班/船名:", "vessel", 4)
        self.addr1_var = add_mdac_config_row(mdac_config_frame, "住宿地址1:", "address1", 5)
        self.addr2_var = add_mdac_config_row(mdac_config_frame, "住宿地址2:", "address2", 6)
        self.postcode_var = add_mdac_config_row(mdac_config_frame, "邮编:", "postcode", 7)

        self.test_mode_var = tk.BooleanVar(value=self.config.get("test_mode", True))
        test_mode_check = ttk.Checkbutton(mdac_config_frame, text="测试模式 (不提交表单)", variable=self.test_mode_var,
                                          command=self.save_config)
        test_mode_check.grid(row=8, column=0, columnspan=3, sticky=tk.W, pady=5)

        mdac_button_frame = ttk.Frame(self.mdac_tab)
        mdac_button_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        self.start_mdac_btn = ttk.Button(mdac_button_frame, text="启动 MDAC 自动化", command=self.start_mdac_automation)
        self.start_mdac_btn.pack(side=tk.LEFT, padx=5, pady=5)

        self.pause_mdac_btn = ttk.Button(mdac_button_frame, text="暂停", command=self.pause_mdac_automation,
                                         state=tk.DISABLED)
        self.pause_mdac_btn.pack(side=tk.LEFT, padx=5, pady=5)

        self.stop_mdac_btn = ttk.Button(mdac_button_frame, text="停止", command=self.stop_mdac_automation,
                                        state=tk.DISABLED)
        self.stop_mdac_btn.pack(side=tk.LEFT, padx=5, pady=5)

        self.save_mdac_config_btn = ttk.Button(mdac_button_frame, text="保存配置", command=self.save_config)
        self.save_mdac_config_btn.pack(side=tk.RIGHT, padx=5, pady=5)

        mdac_log_frame = ttk.LabelFrame(self.mdac_tab, text=" MDAC 运行日志 ", padding=10)
        mdac_log_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.mdac_log_text = scrolledtext.ScrolledText(mdac_log_frame, wrap=tk.WORD, height=15, font=("微软雅黑", 9))
        self.mdac_log_text.pack(fill="both", expand=True)

        # ============================================================
        # Tab 3: Gmail PIN 码自动获取（新增）
        # ============================================================
        self.gmail_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.gmail_tab, text="Gmail PIN 获取")

        gmail_config_frame = ttk.LabelFrame(self.gmail_tab, text=" Gmail 读取设置 ", padding=15)
        gmail_config_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        # Row 0: Gmail 账号
        ttk.Label(gmail_config_frame, text="Gmail 账号:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.gmail_address_var = tk.StringVar(value=self.config.get("gmail_address", ""))
        ttk.Entry(gmail_config_frame, textvariable=self.gmail_address_var, width=40).grid(
            row=0, column=1, sticky=tk.W, padx=10, pady=2)

        # Row 1: Google 16位应用专用密码
        ttk.Label(gmail_config_frame, text="Google 应用专用密码:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.gmail_app_password_var = tk.StringVar(value=self.config.get("gmail_app_password", ""))
        password_entry = ttk.Entry(gmail_config_frame, textvariable=self.gmail_app_password_var, width=40, show="*")
        password_entry.grid(row=1, column=1, sticky=tk.W, padx=10, pady=2)

        # 显示/隐藏密码按钮
        self.show_gmail_pw_var = tk.BooleanVar(value=False)
        show_pw_btn = ttk.Checkbutton(
            gmail_config_frame, text="显示密码", variable=self.show_gmail_pw_var,
            command=lambda: password_entry.configure(show="" if self.show_gmail_pw_var.get() else "*")
        )
        show_pw_btn.grid(row=1, column=2, sticky=tk.W, padx=5, pady=2)

        # Row 2: 检查间隔
        ttk.Label(gmail_config_frame, text="检查间隔 (分钟):").grid(row=2, column=0, sticky=tk.W, pady=2)
        self.gmail_interval_var = tk.StringVar(value=str(self.config.get("gmail_interval", 10)))
        ttk.Entry(gmail_config_frame, textvariable=self.gmail_interval_var, width=10).grid(
            row=2, column=1, sticky=tk.W, padx=10, pady=2)

        # Row 3: Telegram Bot Token (用于报警)
        ttk.Label(gmail_config_frame, text="报警 Bot Token:").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.gmail_telegram_token_var = tk.StringVar(value=self.config.get("gmail_telegram_token", ""))
        ttk.Entry(gmail_config_frame, textvariable=self.gmail_telegram_token_var, width=50).grid(
            row=3, column=1, sticky=tk.W, padx=10, pady=2)

        # Row 4: Telegram Chat ID
        ttk.Label(gmail_config_frame, text="报警接收 Chat ID:").grid(row=4, column=0, sticky=tk.W, pady=2)
        self.gmail_chat_id_var = tk.StringVar(value=self.config.get("gmail_chat_id", ""))
        ttk.Entry(gmail_config_frame, textvariable=self.gmail_chat_id_var, width=50).grid(
            row=4, column=1, sticky=tk.W, padx=10, pady=2)

        # 按钮区
        gmail_button_frame = ttk.Frame(self.gmail_tab)
        gmail_button_frame.pack(fill=tk.X, pady=(0, 10), padx=5)

        self.start_gmail_btn = ttk.Button(gmail_button_frame, text="启动 Gmail 检查", command=self.start_gmail_fetcher)
        self.start_gmail_btn.pack(side=tk.LEFT, padx=5, pady=5)

        self.stop_gmail_btn = ttk.Button(gmail_button_frame, text="停止 Gmail 检查",
                                         command=self.stop_gmail_fetcher, state=tk.DISABLED)
        self.stop_gmail_btn.pack(side=tk.LEFT, padx=5, pady=5)

        self.save_gmail_config_btn = ttk.Button(gmail_button_frame, text="保存配置", command=self.save_config)
        self.save_gmail_config_btn.pack(side=tk.RIGHT, padx=5, pady=5)

        # 日志区
        gmail_log_frame = ttk.LabelFrame(self.gmail_tab, text=" Gmail 运行日志 ", padding=10)
        gmail_log_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.gmail_log_text = scrolledtext.ScrolledText(gmail_log_frame, wrap=tk.WORD, height=15,
                                                         font=("微软雅黑", 9))
        self.gmail_log_text.pack(fill="both", expand=True)

    def browse_excel_file(self, var):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            var.set(file_path)
            self.save_config()

    def process_log_queue(self):
        """
        日志分发到保留的 Tab 1/Tab 3 日志框。
        target_tab: "MDAC" → MDAC框 | "Gmail" → Gmail框
        其他 → 默认进 MDAC 框。
        """
        while not log_queue.empty():
            message, level, target_tab = log_queue.get()
            if target_tab == "MDAC":
                txt = self.mdac_log_text
            elif target_tab == "Gmail":
                txt = self.gmail_log_text
            else:
                txt = self.mdac_log_text  # 默认

            txt.insert(tk.END, f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
            txt.see(tk.END)
        self.root.after(100, self.process_log_queue)

    def start_mdac_automation(self):
        excel_path = self.excel_var.get()
        if not excel_path or not os.path.exists(excel_path):
            messagebox.showerror("错误", "请选择有效的 Excel 文件路径！")
            return
        self.save_config()
        self.excel_manager = ExcelManager(excel_path)
        self.is_mdac_running = True
        self.is_mdac_paused = False
        self.start_mdac_btn.config(state=tk.DISABLED)
        self.pause_mdac_btn.config(state=tk.NORMAL)
        self.stop_mdac_btn.config(state=tk.NORMAL)
        self.mdac_thread = threading.Thread(target=self._run_mdac_automation, daemon=True)
        self.mdac_thread.start()

    def pause_mdac_automation(self):
        self.is_mdac_paused = not self.is_mdac_paused
        self.pause_mdac_btn.config(text="继续" if self.is_mdac_paused else "暂停")

    def stop_mdac_automation(self):
        self.is_mdac_running = False
        self.start_mdac_btn.config(state=tk.NORMAL)
        self.pause_mdac_btn.config(state=tk.DISABLED)
        self.stop_mdac_btn.config(state=tk.DISABLED)

    def _run_mdac_automation(self):
        log_queue.put("🚀 MDAC 自动化线程已启动，正在初始化浏览器...", target_tab="MDAC")
        try:
            with sync_playwright() as p:
                # 尝试启动浏览器
                try:
                    browser = p.chromium.launch(headless=False, channel="chrome")
                    page = browser.new_page()
                    log_queue.put("✅ 浏览器已成功启动，开始扫描 Excel...", target_tab="MDAC")
                except Exception as e:
                    log_queue.put(f"❌ 浏览器启动失败: {e}。请确保已运行 playwright install chromium", level="ERROR",
                                  target_tab="MDAC")
                    return

                while self.is_mdac_running:
                    if self.is_mdac_paused:
                        time.sleep(1)
                        continue

                    try:
                        with self.excel_manager.lock:
                            workbook = load_workbook(self.excel_var.get(), data_only=True)
                            sheet = workbook.active
                    except Exception as e:
                        log_queue.put(f"❌ 读取 Excel 失败: {e}，5秒后重试...", level="ERROR", target_tab="MDAC")
                        time.sleep(5)
                        continue

                    found_pending = False
                    for row in range(2, sheet.max_row + 1):
                        customer = read_customer_from_excel(sheet, row)
                        # 重点检查：这里必须是 PENDING 才会触发
                        if customer["status"] == "PENDING":
                            found_pending = True
                            log_queue.put(f"🔍 发现待处理客户: {customer['name']} (行 {row})，准备填表...",
                                          target_tab="MDAC")
                            process_registration(page, self.excel_manager, row, customer, self.config,
                                                 lambda msg, lvl="INFO": log_queue.put(msg, lvl, "MDAC"))
                            break

                    if not found_pending:
                        # 如果没找到，每 10 秒在日志里刷一下，让你知道它还活着
                        # log_queue.put("😴 未发现 PENDING 状态的客户，10秒后重新扫描...", target_tab="MDAC")
                        time.sleep(10)

            log_queue.put("🛑 MDAC 自动化已停止。", target_tab="MDAC")
        except Exception as e:
            log_queue.put(f"❌ MDAC 线程发生严重错误: {e}", level="ERROR", target_tab="MDAC")
        finally:
            self.stop_mdac_automation()

    # ================================================================
    # Tab 3: Gmail PIN 码获取 启动 / 停止
    # ================================================================
    def start_gmail_fetcher(self):
        """校验配置后启动 Gmail PIN 获取"""
        excel_path = self.excel_var.get()
        gmail_addr = self.gmail_address_var.get().strip()
        gmail_pw = self.gmail_app_password_var.get().strip()
        telegram_token = self.gmail_telegram_token_var.get().strip()
        chat_id = self.gmail_chat_id_var.get().strip()

        # 校验必填项
        if not gmail_addr:
            messagebox.showerror("错误", "请填写 Gmail 账号！")
            return
        if not gmail_pw:
            messagebox.showerror("错误", "请填写 Google 应用专用密码！")
            return
        if not os.path.exists(excel_path):
            messagebox.showerror("错误", "请选择有效的 Excel 文件路径！")
            return

        # 间隔校验
        try:
            interval = int(self.gmail_interval_var.get())
            if interval < 1:
                interval = 1
        except ValueError:
            interval = 10
            self.gmail_interval_var.set("10")

        # 创建 ExcelManager（如果还没创建）
        if self.excel_manager is None or self.excel_manager.file_path != excel_path:
            self.excel_manager = ExcelManager(excel_path)

        # 创建 GmailPINFetcher
        self.gmail_fetcher = GmailPINFetcher(
            email_addr=gmail_addr,
            app_password=gmail_pw,
            excel_manager=self.excel_manager,
            telegram_token=telegram_token,
            chat_id=chat_id,
            interval_minutes=interval,
            log_queue=log_queue,
        )

        self.is_gmail_running = True
        self.start_gmail_btn.config(state=tk.DISABLED)
        self.stop_gmail_btn.config(state=tk.NORMAL)
        self.gmail_fetcher.start()

    def stop_gmail_fetcher(self):
        """停止 Gmail PIN 获取"""
        if self.gmail_fetcher:
            self.gmail_fetcher.stop()
        self.is_gmail_running = False
        self.start_gmail_btn.config(state=tk.NORMAL)
        self.stop_gmail_btn.config(state=tk.DISABLED)


if __name__ == "__main__":
    root = tk.Tk()
    app = MDACApp(root)
    root.mainloop()
